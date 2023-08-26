from django.http import HttpResponse
from django.shortcuts import redirect, render
import pandas as pd
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from django.contrib.auth.decorators import login_required
from django.contrib import auth
from django.contrib.auth import logout
from django.contrib.auth.models import User
from openpyxl import load_workbook

from calc.models import ProductivityData
from django.conf import settings
from django.http import FileResponse

# Create your views here.

def calculate_biomass(volume_ha):
    PAHa = -16.776 + 0.5283 * volume_ha

    CHa = 0.8962 + 0.0283 * volume_ha

    LHa = -22.036 + 0.4688 * volume_ha

    return PAHa, CHa, LHa

def calculate_phosphorus_content(volume_ha):
    PAHa, CHa, LHa = calculate_biomass(volume_ha)

    PPAHa = 3.3845 + 0.1145 * PAHa

    PCHa = 0.7218 + 0.2227 * (LHa + CHa)

    PLHa = 0.4465 + 0.0568 * (LHa + CHa)

    return PPAHa, PCHa, PLHa

def login(request):
    if request.method == "GET":
        return render(request, 'login.html')
    elif request.method == "POST":
        username = request.POST.get('username')

        user, created = User.objects.get_or_create(username=username)

        if created:
            print("Novo usuário criado: {user.username}")

        auth.login(request, user)
        print("Usuário autenticado: {user.username}")
        return redirect('home')

def logooff(request):
    logout(request)

    return render(request, 'login.html')  



def home(request):
    PPAHa, PCHa, PLHa = None, None, None
    
    results = []
    
    if request.method == "POST":
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            with open('temp_excel.xlsx', 'wb') as destination:
                for chunk in excel_file.chunks():
                    destination.write(chunk)

            df = pd.read_excel('temp_excel.xlsx')

            wb = load_workbook('temp_excel.xlsx')
            ws = wb.active

            ws['E1'] = 'PPAHa'
            ws['F1'] = 'PCHa'
            ws['G1'] = 'PLHa'

            for index, row in df.iterrows():
                volume_ha = row['volume_ha']
                PAHa, CHa, LHa = calculate_biomass(volume_ha)
                PPAHa, PCHa, PLHa = calculate_phosphorus_content(volume_ha)
                results.append({
                    'volume_ha': volume_ha,
                    'PPAHa': PPAHa,
                    'PCHa': PCHa,
                    'PLHa': PLHa,
                })
                ws.cell(row=index+2, column=5, value=PPAHa)
                ws.cell(row=index+2, column=6, value=PCHa)
                ws.cell(row=index+2, column=7, value=PLHa)

            os.remove('temp_excel.xlsx')
            
            plan_excel_folder = os.path.join(settings.BASE_DIR, 'plan_excel')
            if not os.path.exists(plan_excel_folder):
                os.makedirs(plan_excel_folder)
            
            resultado_path = os.path.join(plan_excel_folder, 'resultado.xlsx')
            
            wb.save(resultado_path) 

    context = {'results': results}
    return render(request, 'home.html', context)

def download_excel(request):
    resultado_path = os.path.join(settings.BASE_DIR, 'plan_excel', 'resultado.xlsx')
    response = FileResponse(open(resultado_path, 'rb'))
    response['Content-Disposition'] = 'attachment; filename="resultado.xlsx"'
    return response